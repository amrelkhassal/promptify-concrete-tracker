from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Callable, Optional
from uuid import UUID

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.domain.models import FieldType
from app.repos import cache as cache_repo
from app.repos import configs as configs_repo
from app.repos import datasets as datasets_repo
from app.repos.db import (
    ConfigVersion,
    Dataset,
    DatasetDocument,
    Evaluation,
    EvaluationDocument,
    session_scope,
)
from app.services import blob as blob_svc
from app.services.compare import (
    STATUS_GT_EMPTY,
    STATUS_MATCH,
    STATUS_MISMATCH,
    STATUS_MISSING,
    compare_field,
)
from app.services.runner import RunRequest, run_single

log = logging.getLogger(__name__)

# ── Excel colour palette ──────────────────────────────────────────────────────

_FILL = {
    STATUS_MATCH: PatternFill("solid", fgColor="C6EFCE"),
    STATUS_MISMATCH: PatternFill("solid", fgColor="FFC7CE"),
    STATUS_MISSING: PatternFill("solid", fgColor="FFEB9C"),
    STATUS_GT_EMPTY: PatternFill("solid", fgColor="D9D9D9"),
    "header": PatternFill("solid", fgColor="2F5496"),
    "subheader": PatternFill("solid", fgColor="4472C4"),
    "good": PatternFill("solid", fgColor="375623"),
    "ok": PatternFill("solid", fgColor="375623"),
    "bad": PatternFill("solid", fgColor="9C0006"),
}
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD = Font(bold=True, size=11)
_CENTER = Alignment(horizontal="center", vertical="center")


def _pct_fill(pct: Optional[float]) -> PatternFill:
    if pct is None:
        return _FILL[STATUS_GT_EMPTY]
    if pct >= 85:
        return PatternFill("solid", fgColor="C6EFCE")
    if pct >= 60:
        return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="FFC7CE")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _field_type_map(version: ConfigVersion) -> dict[str, FieldType]:
    return {f["name"]: FieldType(f.get("type", "string")) for f in (version.fields or [])}


def _compute_accuracy(statuses: list[str]) -> Optional[float]:
    evaluated = [s for s in statuses if s != STATUS_GT_EMPTY]
    if not evaluated:
        return None
    matches = sum(1 for s in evaluated if s == STATUS_MATCH)
    return round(matches / len(evaluated) * 100, 1)


# ── Run a batch evaluation ────────────────────────────────────────────────────

def run_evaluation(
    eval_id: UUID,
    *,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    """
    Execute a batch evaluation. Runs each document through the provider
    (cache-aware), compares extracted fields to ground truth, persists results,
    and updates the evaluation summary.

    on_progress(current, total, doc_key) — optional callback for UI progress.
    """
    with session_scope() as session:
        evaluation = session.get(Evaluation, eval_id)
        if evaluation is None:
            raise ValueError(f"Evaluation {eval_id} not found")

        version = session.get(ConfigVersion, evaluation.config_version_id)
        if version is None:
            raise ValueError("ConfigVersion not found")

        documents = datasets_repo.get_documents(session, evaluation.dataset_id)
        if not documents:
            evaluation.status = "failed"
            evaluation.finished_at = datetime.now(timezone.utc)
            return

        field_type_map = _field_type_map(version)
        field_names = list(field_type_map.keys())

        # per-field status lists for summary
        field_statuses_agg: dict[str, list[str]] = defaultdict(list)
        worksite_statuses: dict[str, list[str]] = defaultdict(list)
        total = len(documents)

        for i, doc in enumerate(documents):
            if on_progress:
                on_progress(i, total, doc.doc_key)

            file_bytes: Optional[bytes] = None
            error: Optional[str] = None
            extraction: Optional[dict[str, Any]] = None
            latency_ms: Optional[int] = None
            field_statuses: dict[str, str] = {}

            try:
                file_bytes = blob_svc.download_from_path(doc.blob_path)
            except Exception as e:
                error = f"Blob download failed: {e}"
                log.warning("Eval %s: blob error for %s: %s", eval_id, doc.doc_key, e)

            if file_bytes is not None:
                try:
                    from app.domain.models import FieldSpec, ProviderName
                    req = RunRequest(
                        file_bytes=file_bytes,
                        filename=doc.blob_path.rsplit("/", 1)[-1],
                        prompt=version.prompt,
                        fields=[FieldSpec(**f) for f in (version.fields or [])],
                        provider=ProviderName(version.provider),
                        provider_options=version.provider_options or {},
                        skip_cache_lookup=False,
                    )
                    result = run_single(session, req)
                    extraction = result.fields
                    latency_ms = result.latency_ms
                except Exception as e:
                    error = f"OCR failed: {e}"
                    log.warning("Eval %s: OCR error for %s: %s", eval_id, doc.doc_key, e)

            gt = doc.ground_truth or {}
            for field_name in field_names:
                ftype = field_type_map[field_name]
                status = compare_field(
                    gt.get(field_name),
                    (extraction or {}).get(field_name),
                    ftype,
                    name=field_name,
                )
                field_statuses[field_name] = status
                field_statuses_agg[field_name].append(status)

            worksite = (doc.doc_metadata or {}).get("worksite")
            if worksite:
                all_statuses = list(field_statuses.values())
                worksite_statuses[worksite].extend(all_statuses)

            # Upsert EvaluationDocument
            existing_ed = session.execute(
                select(EvaluationDocument).where(
                    EvaluationDocument.evaluation_id == eval_id,
                    EvaluationDocument.document_id == doc.id,
                )
            ).scalar_one_or_none()

            if existing_ed:
                existing_ed.extraction = extraction
                existing_ed.field_statuses = field_statuses
                existing_ed.latency_ms = latency_ms
                existing_ed.error = error
            else:
                session.add(EvaluationDocument(
                    evaluation_id=eval_id,
                    document_id=doc.id,
                    extraction=extraction,
                    field_statuses=field_statuses,
                    latency_ms=latency_ms,
                    error=error,
                ))

        # ── Compute summary ──────────────────────────────────────────────────
        per_field: dict[str, Any] = {}
        for fn, statuses in field_statuses_agg.items():
            counts = {s: statuses.count(s) for s in (STATUS_MATCH, STATUS_MISMATCH, STATUS_MISSING, STATUS_GT_EMPTY)}
            per_field[fn] = {**counts, "accuracy": _compute_accuracy(statuses)}

        per_worksite: dict[str, Any] = {}
        for ws_name, statuses in worksite_statuses.items():
            per_worksite[ws_name] = {"accuracy": _compute_accuracy(statuses), "count": len(statuses)}

        all_statuses_flat = [s for sl in field_statuses_agg.values() for s in sl]
        overall = _compute_accuracy(all_statuses_flat)

        evaluation.summary = {
            "overall_accuracy": overall,
            "per_field": per_field,
            "per_worksite": per_worksite,
            "doc_count": total,
        }
        evaluation.status = "done"
        evaluation.finished_at = datetime.now(timezone.utc)

        if on_progress:
            on_progress(total, total, "done")


# ── Load results for display ──────────────────────────────────────────────────

def get_evaluation_detail(session: Session, eval_id: UUID) -> list[dict[str, Any]]:
    """Return one dict per document: doc_key, worksite, ground_truth, extraction, field_statuses."""
    docs = session.execute(
        select(DatasetDocument, EvaluationDocument)
        .join(EvaluationDocument, EvaluationDocument.document_id == DatasetDocument.id)
        .where(EvaluationDocument.evaluation_id == eval_id)
        .order_by(DatasetDocument.doc_key)
    ).all()
    return [
        {
            "doc_key": dd.doc_key,
            "worksite": (dd.doc_metadata or {}).get("worksite"),
            "ground_truth": dd.ground_truth or {},
            "extraction": ed.extraction or {},
            "field_statuses": ed.field_statuses or {},
            "latency_ms": ed.latency_ms,
            "error": ed.error,
        }
        for dd, ed in docs
    ]


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(session: Session, eval_id: UUID) -> bytes:
    evaluation = session.get(Evaluation, eval_id)
    if evaluation is None:
        raise ValueError("Evaluation not found")

    version = session.get(ConfigVersion, evaluation.config_version_id)
    dataset = session.get(Dataset, evaluation.dataset_id)
    config_name = version.config.name if version else "?"
    summary = evaluation.summary or {}
    per_field = summary.get("per_field", {})
    per_worksite = summary.get("per_worksite", {})
    detail = get_evaluation_detail(session, eval_id)
    field_names = list(per_field.keys())

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.append([f"Config: {config_name} v{version.version if version else '?'}"])
    ws_sum.append([f"Dataset: {dataset.name if dataset else '?'}"])
    ws_sum.append([f"Overall accuracy: {summary.get('overall_accuracy', '?')}%"])
    ws_sum.append([f"Documents: {summary.get('doc_count', '?')}"])
    ws_sum.append([])

    ws_sum.append(["Field", "Accuracy %", "MATCH", "MISMATCH", "MISSING", "GT_EMPTY"])
    header_row = ws_sum.max_row
    for col, val in enumerate(["Field", "Accuracy %", "MATCH", "MISMATCH", "MISSING", "GT_EMPTY"], start=1):
        cell = ws_sum.cell(row=header_row, column=col)
        cell.fill = _FILL["header"]
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for fn in field_names:
        fd = per_field[fn]
        pct = fd.get("accuracy")
        row = [fn, f"{pct}%" if pct is not None else "—",
               fd.get(STATUS_MATCH, 0), fd.get(STATUS_MISMATCH, 0),
               fd.get(STATUS_MISSING, 0), fd.get(STATUS_GT_EMPTY, 0)]
        ws_sum.append(row)
        pct_cell = ws_sum.cell(row=ws_sum.max_row, column=2)
        pct_cell.fill = _pct_fill(pct)

    ws_sum.column_dimensions["A"].width = 28
    for col in "BCDEF":
        ws_sum.column_dimensions[col].width = 14

    # ── Sheet 2: By Worksite ──────────────────────────────────────────────────
    if per_worksite:
        ws_ws = wb.create_sheet("By Worksite")
        ws_ws.append(["Worksite", "Accuracy %", "Evaluated fields"])
        for col, val in enumerate(["Worksite", "Accuracy %", "Evaluated fields"], start=1):
            cell = ws_ws.cell(row=1, column=col)
            cell.fill = _FILL["header"]
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
        for ws_name, wd in sorted(per_worksite.items()):
            pct = wd.get("accuracy")
            ws_ws.append([ws_name, f"{pct}%" if pct is not None else "—", wd.get("count", 0)])
            ws_ws.cell(row=ws_ws.max_row, column=2).fill = _pct_fill(pct)
        ws_ws.column_dimensions["A"].width = 22
        ws_ws.column_dimensions["B"].width = 14
        ws_ws.column_dimensions["C"].width = 18

    # ── Sheet 3: Details ──────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Details")
    # Two header rows: doc info columns, then alternating GT/Extracted per field
    header1 = ["Document", "Worksite", "Error"] + [fn for fn in field_names for _ in ("GT", "Extracted")]
    header2 = ["", "", ""] + [label for fn in field_names for label in ("GT", "Extracted")]
    ws_det.append(header1)
    ws_det.append(header2)

    # Style double header
    for row_idx in (1, 2):
        for col_idx in range(1, len(header1) + 1):
            c = ws_det.cell(row=row_idx, column=col_idx)
            c.fill = _FILL["header"]
            c.font = _HEADER_FONT
            c.alignment = _CENTER

    for doc in detail:
        row_values: list[Any] = [doc["doc_key"], doc.get("worksite") or "", doc.get("error") or ""]
        fills: list[Optional[PatternFill]] = [None, None, None]
        for fn in field_names:
            gt_val = doc["ground_truth"].get(fn)
            ex_val = doc["extraction"].get(fn)
            status = doc["field_statuses"].get(fn, "")
            row_values += [gt_val if gt_val is not None else "—", ex_val if ex_val is not None else "—"]
            fill = _FILL.get(status)
            fills += [fill, fill]

        ws_det.append(row_values)
        for col_idx, fill in enumerate(fills, start=1):
            if fill is not None:
                ws_det.cell(row=ws_det.max_row, column=col_idx).fill = fill

    ws_det.column_dimensions["A"].width = 28
    ws_det.column_dimensions["B"].width = 18
    ws_det.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
