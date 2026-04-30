import io
import zipfile

import pandas as pd
import streamlit as st

from app.core.auth import get_current_user
from app.core.config import get_settings
from app.core.seed import GT_COLUMN_MAP, WORKSITE_MAP, _coerce_gt_value, _normalize_key
from app.repos import datasets as datasets_repo
from app.repos.db import session_scope
from app.services import blob as blob_svc


st.set_page_config(page_title="Datasets · Promptify", layout="wide")
st.title("Datasets")
st.caption("Ground-truth labelled document sets used for batch evaluation.")


# ── Dataset list ──────────────────────────────────────────────────────────────

with session_scope() as s:
    all_datasets = datasets_repo.list_datasets(s)

if all_datasets:
    rows = [
        {
            "Name": d["dataset"].name,
            "Documents": d["doc_count"],
            "Description": d["dataset"].description or "—",
            "Created by": d["dataset"].created_by,
            "Created at": d["dataset"].created_at.strftime("%Y-%m-%d"),
        }
        for d in all_datasets
    ]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
else:
    st.info("No datasets yet. Upload one below or let the seed run on startup.")

st.divider()


# ── Inspect a dataset ─────────────────────────────────────────────────────────

if all_datasets:
    ds_names = [d["dataset"].name for d in all_datasets]
    selected = st.selectbox("Inspect dataset", ds_names)
    sel_ds = next(d["dataset"] for d in all_datasets if d["dataset"].name == selected)

    with session_scope() as s:
        docs = datasets_repo.get_documents(s, sel_ds.id)

    if docs:
        doc_rows = [
            {
                "doc_key": d.doc_key,
                "worksite": (d.doc_metadata or {}).get("worksite", "—"),
                "mime_type": d.mime_type,
                "blob_path": d.blob_path,
                "gt_fields": len([v for v in (d.ground_truth or {}).values() if v is not None]),
            }
            for d in docs
        ]
        st.dataframe(pd.DataFrame(doc_rows), use_container_width=True, hide_index=True)
    else:
        st.warning("No documents in this dataset.")

    st.divider()


# ── Upload new dataset ────────────────────────────────────────────────────────

with st.expander("➕ Upload new dataset", expanded=not bool(all_datasets)):
    st.markdown(
        """
Upload a **ZIP file** containing the BL documents (PDF/JPEG/PNG) and an **Excel/CSV file**
with ground-truth field values.

**Excel format** — same column headers as `Tests.xlsx`:

| Nom du fichier | Numéro du bon | Quantité livrée | … |
|---|---|---|---|
| FR01 - 12345 | 12345 | 7.5 | … |

The `Nom du fichier` column must match the file stem (without extension).
"""
    )

    new_ds_name = st.text_input("Dataset name", placeholder="e.g. cyprus-2025")
    new_ds_desc = st.text_input("Description (optional)")
    zip_file = st.file_uploader("Documents ZIP", type=["zip"], key="ds_zip")
    gt_file = st.file_uploader("Ground-truth Excel or CSV", type=["xlsx", "csv"], key="ds_gt")

    if st.button("Upload dataset", disabled=not (new_ds_name.strip() and zip_file and gt_file)):
        try:
            blob_svc.ensure_containers()

            # ── Parse ground truth ────────────────────────────────────────────
            import openpyxl, csv
            gt_data = gt_file.read()
            if gt_file.name.endswith(".csv"):
                reader = csv.DictReader(io.StringIO(gt_data.decode("utf-8-sig")))
                rows_raw = list(reader)
                gt_rows = {
                    str(r.get("Nom du fichier", "")).strip(): r for r in rows_raw
                }
            else:
                wb = openpyxl.load_workbook(io.BytesIO(gt_data))
                ws = wb.active
                headers = [c.value for c in ws[1]]
                gt_rows = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rd = dict(zip(headers, row))
                    stem = str(rd.get("Nom du fichier") or "").strip()
                    if stem:
                        gt_rows[stem] = rd

            gt_by_norm = {_normalize_key(k): (k, v) for k, v in gt_rows.items()}

            # ── Create dataset ────────────────────────────────────────────────
            with session_scope() as s:
                ds = datasets_repo.create_dataset(
                    s,
                    name=new_ds_name.strip(),
                    description=new_ds_desc.strip() or None,
                    created_by=get_current_user(),
                )
                ds_id = ds.id

            # ── Extract ZIP and upload each file ──────────────────────────────
            with zipfile.ZipFile(io.BytesIO(zip_file.read())) as zf:
                names = [n for n in zf.namelist() if not n.endswith("/") and "." in n.rsplit("/", 1)[-1]]
                bar = st.progress(0, text="Uploading…")
                inserted = skipped = 0
                for idx, zname in enumerate(names):
                    filename = zname.rsplit("/", 1)[-1]
                    stem = filename.rsplit(".", 1)[0]
                    norm = _normalize_key(stem)
                    match = gt_by_norm.get(norm)
                    if match is None:
                        skipped += 1
                        continue
                    _, row_dict = match

                    mime = blob_svc.mime_for_filename(filename)
                    blob_name = f"{new_ds_name.strip()}/{filename}"
                    data = zf.read(zname)
                    blob_svc.upload(get_settings().blob_container_datasets, blob_name, data, mime)
                    blob_path = f"{get_settings().blob_container_datasets}/{blob_name}"

                    gt: dict = {}
                    for col_name, field_name in GT_COLUMN_MAP.items():
                        raw = row_dict.get(col_name)
                        gt[field_name] = _coerce_gt_value(field_name, raw)

                    with session_scope() as s:
                        datasets_repo.add_document(
                            s,
                            dataset_id=ds_id,
                            doc_key=stem,
                            blob_path=blob_path,
                            mime_type=mime,
                            ground_truth=gt,
                            metadata=None,
                        )
                    inserted += 1
                    bar.progress((idx + 1) / len(names), text=f"Uploading {filename}…")

            st.success(f"Dataset **{new_ds_name}** created with {inserted} documents ({skipped} skipped — no GT match).")
            st.rerun()
        except Exception as e:
            st.exception(e)
