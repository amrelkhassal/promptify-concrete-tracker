import logging
import re
from pathlib import Path
from typing import Any, Optional

from app.core.auth import get_current_user
from app.domain.defaults import DEFAULT_FIELDS, DEFAULT_PROMPT
from app.domain.models import ProviderName
from app.repos import configs as configs_repo
from app.repos import datasets as datasets_repo
from app.repos.db import session_scope
from app.services import blob as blob_svc
from app.services.normalize import normalize_date, normalize_time


log = logging.getLogger(__name__)

DEFAULT_CONFIG_NAME = "Default"
SEED_DATASET_NAME = "seed-50bls"

# Maps Tests.xlsx column headers → field names in our schema
GT_COLUMN_MAP = {
    "Numéro du bon": "deliveryNumber",
    "Quantité livrée": "quantity",
    "Formule commerciale": "typeBeton",
    "Date de livraison": "dateLivraison",
    "Date/heure de 1ère gâchée": "heurePremiereGachee",
    "Date/heure d'arrivée sur chantier": "heureArriveeChantier",
    "Date/heure de début de déchargement": "heureDebutDechargement",
    "Date/heure de fin de déchargement": "heureFinDechargement",
    "Date/heure de départ chantier": "heureDepartChantier",
    "Addition": "additive",
    "Adjuvant": "adjuvant",
    "Poids carbone": "carboneWeight",
    "Classe d'exposition": "classeExposition",
    "Classe de résistance": "classeResistance",
    "Classe de chlorures": "classeChlorures",
    "Classe de consistance": "classeConsistance",
    "Famille de ciment": "familleCiment",
    "Diamètre de granulat": "diametreGranulat",
    "Désignation des bétons": "designation",
}

WORKSITE_MAP = {
    "10 BLs Saint-denis": "Saint-Denis",
    "10 BLs Viroflay": "Viroflay",
    "10 BLs Amiens": "Amiens",
    "10 BLs Arabie Saoudite": "Arabie Saoudite",
    "10 BLs Chypre": "Chypre",
}


def _normalize_key(s: str) -> str:
    """Collapse all whitespace for fuzzy file-stem ↔ GT-key matching."""
    return re.sub(r"\s+", "", s).upper()


def _coerce_gt_value(field_name: str, raw: Any) -> Optional[str]:
    """Convert Excel cell value (datetime/time/float/str) to a normalized string."""
    import datetime as dt
    if raw is None:
        return None
    if field_name in ("dateLivraison",):
        return normalize_date(raw)
    if field_name in (
        "heurePremiereGachee", "heureArriveeChantier",
        "heureDebutDechargement", "heureFinDechargement", "heureDepartChantier",
    ):
        return normalize_time(raw)
    if isinstance(raw, dt.datetime):
        return normalize_date(raw)
    if isinstance(raw, dt.time):
        return normalize_time(raw)
    if isinstance(raw, float):
        return str(raw)
    return str(raw).strip() if str(raw).strip() else None


def seed_default_config() -> None:
    """Idempotently create the seed 'Default' config."""
    with session_scope() as session:
        existing = configs_repo.get_config_by_name(session, DEFAULT_CONFIG_NAME)
        if existing is not None:
            log.info("Seed: %r already exists (id=%s), skipping.", DEFAULT_CONFIG_NAME, existing.id)
            return

        created = configs_repo.create_config(
            session,
            name=DEFAULT_CONFIG_NAME,
            description="Bundled starting prompt + 22-field schema.",
            prompt=DEFAULT_PROMPT,
            fields=DEFAULT_FIELDS,
            provider=ProviderName.MISTRAL_2505,
            provider_options={},
            notes="Seed v1 — first checkpoint, do not delete.",
            created_by=get_current_user(),
        )
        log.info("Seed: created %r v%d (id=%s)", created.config_name, created.version, created.config_id)


def seed_dataset(batch_test_dir: Optional[Path] = None) -> None:
    """
    Idempotently ingest the 50-BL corpus from the BatchTest directory into
    Blob Storage + the database as the 'seed-50bls' dataset.

    batch_test_dir defaults to the bundled data/seed/BatchTest path.
    """
    if batch_test_dir is None:
        # Bundled in the Docker image at /app/data/seed/BatchTest
        batch_test_dir = Path(__file__).parent.parent.parent / "data" / "seed" / "BatchTest"

    if not batch_test_dir.exists():
        log.warning("Seed dataset: BatchTest dir not found at %s — skipping.", batch_test_dir)
        return

    xlsx_path = batch_test_dir / "Tests.xlsx"
    if not xlsx_path.exists():
        log.warning("Seed dataset: Tests.xlsx not found — skipping.")
        return

    with session_scope() as session:
        existing = datasets_repo.get_dataset_by_name(session, SEED_DATASET_NAME)
        if existing is not None:
            log.info("Seed dataset: %r already exists, skipping.", SEED_DATASET_NAME)
            return

    # ── Load ground truth ─────────────────────────────────────────────────────
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    gt_by_norm_key: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        file_stem = str(row_dict.get("Nom du fichier") or "").strip()
        if not file_stem:
            continue
        norm = _normalize_key(file_stem)
        gt: dict[str, Any] = {}
        for col_name, field_name in GT_COLUMN_MAP.items():
            raw = row_dict.get(col_name)
            gt[field_name] = _coerce_gt_value(field_name, raw)
        gt_by_norm_key[norm] = gt

    log.info("Seed dataset: loaded %d GT rows from Tests.xlsx.", len(gt_by_norm_key))

    # ── Ensure Blob containers exist ──────────────────────────────────────────
    try:
        blob_svc.ensure_containers()
    except Exception as e:
        log.warning("Seed dataset: could not ensure Blob containers: %s", e)

    # ── Upload files + insert DB rows ─────────────────────────────────────────
    with session_scope() as session:
        ds = datasets_repo.create_dataset(
            session,
            name=SEED_DATASET_NAME,
            description="50 concrete delivery notes across 5 worksites (Saint-Denis, Viroflay, Amiens, Arabie Saoudite, Chypre).",
            created_by=get_current_user(),
        )
        inserted = 0
        skipped = 0

        for folder_name, worksite in WORKSITE_MAP.items():
            folder = batch_test_dir / folder_name
            if not folder.exists():
                log.warning("Seed dataset: folder not found: %s", folder)
                continue

            for file_path in sorted(folder.iterdir()):
                if not file_path.is_file():
                    continue
                stem = file_path.stem
                norm = _normalize_key(stem)
                gt = gt_by_norm_key.get(norm)
                if gt is None:
                    log.warning("Seed dataset: no GT row for %s (norm=%s)", stem, norm)
                    skipped += 1
                    continue

                mime = blob_svc.mime_for_filename(file_path.name)
                blob_name = f"seed-50bls/{worksite}/{file_path.name}"

                try:
                    data = file_path.read_bytes()
                    blob_svc.upload(get_settings().blob_container_datasets, blob_name, data, mime)
                except Exception as e:
                    log.warning("Seed dataset: Blob upload failed for %s: %s — storing path reference only", file_path.name, e)

                blob_path = f"{get_settings().blob_container_datasets}/{blob_name}"
                datasets_repo.add_document(
                    session,
                    dataset_id=ds.id,
                    doc_key=stem,
                    blob_path=blob_path,
                    mime_type=mime,
                    ground_truth=gt,
                    metadata={"worksite": worksite},
                )
                inserted += 1

        log.info(
            "Seed dataset: created %r with %d documents (%d skipped, no GT).",
            SEED_DATASET_NAME, inserted, skipped,
        )


from app.core.config import get_settings
